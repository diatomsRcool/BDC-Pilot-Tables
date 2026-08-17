"""
03_build_table.py

Join phv metadata with counts, aggregate into the before table,
and export to Excel.

Table structure:
  Rows:    Conditions | Drug Exposures | Procedures | SdohObservations |
           <one row per continuous variable concept, alphabetical> | Total
  Columns: For each cohort: "n vars" and "n data pts"
           Final two columns: Total n vars | Total n data pts

Usage:
    python scripts/03_build_table.py

Inputs:
    data/phv_by_cohort_class.csv   (from 01_parse_yaml_phvs.py)
    data/phv_counts.csv            (from 02_fetch_counts.py)

Output:
    tables/before_table.xlsx
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# Constants
# --------------------------------------------------------------------------

PHV_CSV = Path("data/phv_by_cohort_class.csv")
COUNTS_CSV = Path("data/phv_counts.csv")
OUTPUT_XLSX = Path("tables/before_table.xlsx")

# Categorical rows — always appear first and are collapsed by BDCHM class
CATEGORICAL_ROWS = [
    "Conditions",
    "Drug Exposures",
    "Procedures",
]

# Cohort column order (alphabetical)
COHORT_ORDER = ["ARIC", "CARDIA", "CHS", "COPDGene", "FHS", "HCHS", "JHS", "MESA", "WHI"]

# Variable files that should always appear as continuous concept rows,
# even if some cohorts classify them under a categorical BDCHM class.
FORCE_CONTINUOUS = {"edu_lvl", "fam_income"}

# Variable files that should always appear in the Procedures categorical row,
# even if some cohorts classify them under a different BDCHM class.
FORCE_PROCEDURES = {"pacem_stat"}

# Variable files excluded from the Conditions row only.
# They may still appear in other categorical rows (Drug Exposures, Procedures)
# if other cohorts classify them differently.
EXCLUDE_FROM_CONDITIONS = {"chr_bronchitis", "emphysema", "hypert_trt", "hist_cor_bypg"}

# Concept merges: map non-canonical variable_file names to their canonical name.
# Both variable files represent the same underlying concept across cohorts.
CONCEPT_MERGE = {
    "alcohol":            "alcohol_servings",
    "basophil_ct":        "basophil_ncnc_bld",
    "eosinophil_ct":      "eosinophil_ncnc_bld",
    "fasting_blood_gluc": "fast_gluc_bld",
    "hrt_rt":             "hrtrt",
    "insulin_in_blood":   "insulin_blood",
    "lymphocyte_ct":      "lympho_ct",
    "mn_art_press":       "mean_art_press",
    "monocyte_ct":        "monocyte_ncnc_bld",
    "neutro_ct":          "neutrophil_ct",
    "rbc":                "rdbld_ct",
}


# --------------------------------------------------------------------------
# Build the pivot table
# --------------------------------------------------------------------------

def build_table(phv_df: pd.DataFrame, counts_df: pd.DataFrame):
    """
    Join phv metadata with counts, then pivot to the publication table format.

    Returns (table_df, n_categorical_rows, continuous_concepts).
    """
    # Merge: left join so we keep all phvs even if count is missing
    merged = phv_df.merge(
        counts_df[["phv", "phs", "n", "count_available"]],
        on=["phv", "phs"],
        how="left",
    )
    merged["n"] = merged["n"].fillna(0).astype(int)
    merged["count_available"] = merged["count_available"].fillna(False)

    missing_counts = merged[~merged["count_available"]]
    if len(missing_counts) > 0:
        print(f"WARNING: {len(missing_counts)} phvs have no count (treated as 0):")
        print(missing_counts[["cohort", "row_category", "phv"]].to_string(index=False))

    # Force certain variable files to continuous regardless of BDCHM class
    merged.loc[merged["variable_file"].isin(FORCE_CONTINUOUS), "row_category"] = "Continuous variables"

    # Force certain variable files to Procedures regardless of BDCHM class
    merged.loc[merged["variable_file"].isin(FORCE_PROCEDURES), "row_category"] = "Procedures"

    # Drop excluded files from Conditions only
    merged = merged[
        ~(
            (merged["variable_file"].isin(EXCLUDE_FROM_CONDITIONS))
            & (merged["row_category"] == "Conditions")
        )
    ]

    # Apply concept merges: rename non-canonical variable_file names so that
    # cohort-specific synonyms aggregate into a single row.
    merged["variable_file"] = merged["variable_file"].replace(CONCEPT_MERGE)

    # Split into categorical and continuous
    cat_data  = merged[merged["row_category"] != "Continuous variables"]
    cont_data = merged[merged["row_category"] == "Continuous variables"]

    # Deduplicate: a phv may appear in multiple derivation blocks / pht tables.
    # For categorical rows, deduplicate per (cohort, row_category, phv).
    # For continuous rows, deduplicate per (cohort, variable_file, phv) so each
    # concept row gets an independent count of its source phvs.
    cat_deduped  = cat_data.drop_duplicates(subset=["cohort", "row_category", "phv"])
    cont_deduped = cont_data.drop_duplicates(subset=["cohort", "variable_file", "phv"])

    # Aggregate categorical rows by class
    cat_agg = (
        cat_deduped.groupby(["cohort", "row_category"])
        .agg(n_vars=("phv", "nunique"), n_data_pts=("n", "sum"))
        .reset_index()
        .rename(columns={"row_category": "row_label"})
    )

    # Aggregate continuous rows by variable concept (variable_file)
    cont_agg = (
        cont_deduped.groupby(["cohort", "variable_file"])
        .agg(n_vars=("phv", "nunique"), n_data_pts=("n", "sum"))
        .reset_index()
        .rename(columns={"variable_file": "row_label"})
    )

    all_agg = pd.concat([cat_agg, cont_agg], ignore_index=True)

    # Row order: categorical section first, then continuous concepts alphabetically
    cont_concepts = sorted(cont_data["variable_file"].unique())
    row_order = CATEGORICAL_ROWS + cont_concepts

    # Pivot to wide format: one column-pair per cohort
    wide_vars = all_agg.pivot(index="row_label", columns="cohort", values="n_vars").fillna(0).astype(int)
    wide_pts  = all_agg.pivot(index="row_label", columns="cohort", values="n_data_pts").fillna(0).astype(int)

    wide_vars = wide_vars.reindex(index=row_order, columns=COHORT_ORDER, fill_value=0)
    wide_pts  = wide_pts.reindex(index=row_order, columns=COHORT_ORDER, fill_value=0)

    # Interleave cohort columns: ARIC n vars, ARIC n data pts, CARDIA n vars, ...
    data_cols = {}
    for cohort in COHORT_ORDER:
        data_cols[f"{cohort}\nn vars"]     = wide_vars[cohort]
        data_cols[f"{cohort}\nn data pts"] = wide_pts[cohort]

    table = pd.DataFrame(data_cols, index=row_order)

    # Total columns (sum across all cohorts per row)
    table["Total\nn vars"]     = wide_vars.sum(axis=1)
    table["Total\nn data pts"] = wide_pts.sum(axis=1)

    # Total row (sum across all row labels per column)
    # For the total row, sum only the categorical rows + per-concept continuous
    # rows — this may double-count phvs shared across concepts, but that is
    # consistent with how each section is independently counted.
    totals = table.sum(axis=0).rename("Total")
    table = pd.concat([table, totals.to_frame().T])
    table.index.name = "Variable Type"
    table = table.reset_index()

    return table, len(CATEGORICAL_ROWS), cont_concepts


# --------------------------------------------------------------------------
# Excel formatting
# --------------------------------------------------------------------------

HEADER_FILL   = PatternFill("solid", fgColor="2E4057")   # dark blue
CATEG_FILL    = PatternFill("solid", fgColor="D5E8D4")   # light green — categorical section
TOTAL_FILL    = PatternFill("solid", fgColor="D6E4F0")   # light blue — grand total row
SUBTOTAL_FILL = PatternFill("solid", fgColor="EBF5FB")   # very light blue — total columns

HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BOLD_FONT   = Font(bold=True, size=10)
NORMAL_FONT = Font(size=10)


def apply_formatting(ws, n_categorical: int):
    """
    Apply formatting to the worksheet after data is written.

    n_categorical: number of categorical rows (Conditions, Drug Exposures, etc.)
                   These appear in rows 2..(2 + n_categorical - 1) and get a
                   distinct background to separate them from continuous rows.
    """
    max_row = ws.max_row
    max_col = ws.max_column
    total_col_start = max_col - 1  # last two columns are grand totals

    # Row 1: header
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # Data rows (rows 2 .. max_row - 1)
    for row in range(2, max_row):
        is_categorical = row <= (1 + n_categorical)
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(
                horizontal="center" if col > 1 else "left",
                vertical="center",
            )
            if col > 1:
                cell.number_format = "#,##0"

            # Categorical section: bold + green fill
            if is_categorical:
                cell.font = BOLD_FONT
                if col < total_col_start:
                    cell.fill = CATEG_FILL
                else:
                    cell.fill = SUBTOTAL_FILL
            else:
                cell.font = NORMAL_FONT
                if col >= total_col_start:
                    cell.fill = SUBTOTAL_FILL

    # Totals row (last row): bold + blue fill
    for col in range(1, max_col + 1):
        cell = ws.cell(row=max_row, column=col)
        cell.font = BOLD_FONT
        cell.fill = TOTAL_FILL
        cell.alignment = Alignment(horizontal="center" if col > 1 else "left",
                                   vertical="center")
        if col > 1:
            cell.number_format = "#,##0"

    # Total columns (last two) on continuous rows: light blue fill
    for row in range(2 + n_categorical, max_row):
        for col in range(total_col_start, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BOLD_FONT
            cell.fill = SUBTOTAL_FILL
            cell.number_format = "#,##0"

    # Column widths
    ws.column_dimensions[get_column_letter(1)].width = 24  # Variable Type
    for col in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13

    # Row heights
    ws.row_dimensions[1].height = 36  # header
    for row in range(2, max_row + 1):
        ws.row_dimensions[row].height = 18

    # Freeze panes: freeze the header row and Variable Type column
    ws.freeze_panes = "B2"


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def main():
    for path in [PHV_CSV, COUNTS_CSV]:
        if not path.exists():
            sys.exit(f"ERROR: {path} not found. Run earlier scripts first.")

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)

    phv_df    = pd.read_csv(PHV_CSV, dtype=str)
    counts_df = pd.read_csv(COUNTS_CSV, dtype={"n": int, "count_available": bool})

    print(f"Loaded {len(phv_df)} phv rows and {len(counts_df)} count rows")

    table, n_categorical, cont_concepts = build_table(phv_df, counts_df)

    print(f"\nTable dimensions: {len(table)} rows x {len(table.columns)} columns")
    print(f"  Categorical rows: {n_categorical}")
    print(f"  Continuous concept rows: {len(cont_concepts)}")
    print(f"  Total row: 1")

    print("\nBefore table (preview — first 10 rows):")
    print(table.head(10).to_string(index=False))

    # Write to Excel
    table.to_excel(OUTPUT_XLSX, sheet_name="Before", index=False)

    # Apply formatting
    wb = load_workbook(OUTPUT_XLSX)
    ws = wb["Before"]
    apply_formatting(ws, n_categorical)
    wb.save(OUTPUT_XLSX)

    print(f"\nSaved formatted table to {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
